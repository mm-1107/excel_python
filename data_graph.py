#! python
# 滞在期間ごとのUU数のエクセルデータがあった時に
import openpyxl as px
import pprint
import sys
import os.path

# 引数にxlsxファイルパスを指定
args = sys.argv
filename = args[1]

# ファイル読み込み
wb = px.load_workbook(filename)
# シートを指定して読み込み
sheet = wb.get_sheet_by_name('Sheet1')

# sheet.max_rowの値を取っておく
ROW = sheet.max_row
SUM = 0

##################################
# floor(A2/7,1)と同じ
def calculate_floor():
    sheet.cell(row=1,column=3).value = "floor()"
    # c列にfloorの値を入れていく(1日の場合を含める)
    global floor,SUM
    for value in range(2,ROW+1):
        a_cell = sheet.cell(row=value,column=1).value
        floor = a_cell // 7
        # c列にfloorの値を代入(1日を除く)
        sheet.cell(row=value,column=3).value = floor
        # print(floor)
        SUM += sheet.cell(row=value,column=2).value

##################################


##################################
# 週ごとの合計を求める
def calculate_sum():
    sheet.cell(row=1,column=4).value = "週"
    sheet.cell(row=1,column=5).value = "合計(週ごと)"

    # 一週間ごとの合計をE列に入れる
    for value in range(0,floor+1):
        sheet.cell(row=value+2,column=4).value = value

        # E列を初期化
        sheet.cell(row=value+2,column=5).value = 0

        b_cell = sheet.cell(row=2,column=2).value
        c_cell = sheet.cell(row=2,column=3).value

        i=0
        j=0

        # SUMIF(C:C,D2,B:B)と同じ
        while (i < ROW and j<7):
            # floorの値とvalueが同じなら(同週なら)週ごとの合計を求める
            if(c_cell == value):
                j+=1
                sheet.cell(row=value+2,column=5).value += b_cell
            # 次のセルをみる
            i+=1
            b_cell = sheet.cell(row=2+i,column=2).value
            c_cell = sheet.cell(row=2+i,column=3).value

##################################


##################################
# 割合を求める
def calculate_raito():
    sheet.cell(row=1,column=6).value = "割合"
    for value in range(0,floor+1):
        e_cell = sheet.cell(row=value+2,column=5).value
        sheet.cell(row=value+2,column=6).value = 0
        sheet.cell(row=value+2,column=6).value = (e_cell / SUM) *100

##################################


##################################
# 累計割合を求める
def calculate_total_ratio():
    sheet.cell(row=1,column=7).value = "累計割合"
    sheet.cell(row=2,column=7).value = sheet.cell(row=2,column=6).value
    for value in range(0,floor):
        f_cell = sheet.cell(row=value+3,column=6).value
        sheet.cell(row=value+3,column=7).value = f_cell + sheet.cell(row=value+2,column=7).value
        # print(sheet.cell(row=value+3,column=7).value)

##################################


##################################
# グラフをかく
def graph():
    ref_obj_1 = px.chart.Reference(sheet,min_row=2,min_col=5,max_row=floor+2,max_col=5)
    series_obj_1 = px.chart.Series(ref_obj_1,title='UU数合計(週ごと)')
    chart_obj_1 = px.chart.BarChart()
    chart_obj_1.append(series_obj_1)

    ref_obj_2 = px.chart.Reference(sheet,min_row=2,min_col=6,max_row=floor+2,max_col=6)
    series_obj_2 = px.chart.Series(ref_obj_2,title='全体UU数に対する対象UU数の割合')
    chart_obj_2 = px.chart.BarChart()
    chart_obj_2.append(series_obj_2)

    ref_obj_3 = px.chart.Reference(sheet,min_row=2,min_col=7,max_row=floor+2,max_col=7)
    series_obj_3 = px.chart.Series(ref_obj_3,title='累計割合')
    chart_obj_3 = px.chart.BarChart()
    chart_obj_3.append(series_obj_3)

    sheet.add_chart(chart_obj_1)
    sheet.add_chart(chart_obj_2)
    sheet.add_chart(chart_obj_3)

##################################

calculate_floor()
calculate_sum()
calculate_raito()
calculate_total_ratio()
graph()

# 編集したファイルの保存
wb.save(filename)
